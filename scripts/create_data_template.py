#!/usr/bin/env python3
"""
Create an enhanced Excel data template for qualitative data analysis.
This template helps users format their data correctly for upload.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_data_template(output_path):
    """Create a comprehensive data template with formatting and instructions."""

    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    required_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    optional_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    instruction_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1F4E78")
    subtitle_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Instructions
    ws_instructions = wb.create_sheet("📋 Instructions", 0)
    ws_instructions.column_dimensions['A'].width = 90

    instructions = [
        ("Data Template for Qualitative Analysis Tool", title_font, None),
        ("", None, None),
        ("OVERVIEW", subtitle_font, instruction_fill),
        ("This template helps you format your qualitative data for upload and analysis. The tool uses machine learning to automatically discover themes and codes in open-ended text responses.", None, None),
        ("", None, None),
        ("REQUIRED FIELD (must be present)", subtitle_font, required_fill),
        ("• response - The open-ended text response to analyze", None, None),
        ("  - Minimum length: 5 characters (configurable during analysis)", None, None),
        ("  - Type: Text/String", None, None),
        ("  - Example: 'I love the flexibility of remote work'", None, None),
        ("", None, None),
        ("OPTIONAL FIELDS (recommended but not required)", subtitle_font, optional_fill),
        ("• id - Unique identifier for each response", None, None),
        ("  - Type: Number or Text", None, None),
        ("  - Example: 1, 2, 3 or 'RESP001', 'RESP002'", None, None),
        ("", None, None),
        ("• respondent_id - Unique identifier for each participant", None, None),
        ("  - Type: Number or Text", None, None),
        ("  - Example: 'R001', 'R002' or 101, 102", None, None),
        ("", None, None),
        ("• timestamp - Date/time when response was collected", None, None),
        ("  - Format: YYYY-MM-DD or YYYY-MM-DD HH:MM:SS", None, None),
        ("  - Example: '2024-01-15' or '2024-01-15 14:30:00'", None, None),
        ("", None, None),
        ("• topic - Topic or category for the response", None, None),
        ("  - Type: Text", None, None),
        ("  - Example: 'Product Feedback', 'Customer Service'", None, None),
        ("", None, None),
        ("• You can add any other demographic or grouping columns you need (e.g., age, region, department)", None, None),
        ("", None, None),
        ("DATA QUALITY REQUIREMENTS", subtitle_font, instruction_fill),
        ("• Minimum responses: At least 1 response required (30+ recommended for quality results)", None, None),
        ("• Response length: Minimum 5 characters (you can adjust this during analysis)", None, None),
        ("• Text quality: Raw text without special formatting works best", None, None),
        ("• Duplicates: The tool can automatically remove duplicate responses", None, None),
        ("• Missing data: Null/empty responses can be automatically filtered out", None, None),
        ("", None, None),
        ("SUPPORTED FILE FORMATS", subtitle_font, instruction_fill),
        ("• Excel files: .xlsx, .xls", None, None),
        ("• CSV files: .csv", None, None),
        ("• Database connections: SQLite, PostgreSQL", None, None),
        ("", None, None),
        ("HOW TO USE THIS TEMPLATE", subtitle_font, instruction_fill),
        ("1. Go to the '✏️ Data Entry' sheet", None, None),
        ("2. Enter your data starting in row 2 (headers are in row 1)", None, None),
        ("3. At minimum, fill in the 'response' column with your text data", None, None),
        ("4. Add optional columns (id, respondent_id, etc.) if you have them", None, None),
        ("5. Save the file as .xlsx or .csv format", None, None),
        ("6. Upload to the analysis tool", None, None),
        ("", None, None),
        ("ANALYSIS FEATURES", subtitle_font, instruction_fill),
        ("The tool will automatically:", None, None),
        ("• Discover themes and codes in your text data using machine learning", None, None),
        ("• Assign codes to each response with confidence scores", None, None),
        ("• Generate a codebook with definitions and keywords", None, None),
        ("• Create visualizations (word clouds, co-occurrence networks, etc.)", None, None),
        ("• Export results to Excel, CSV, or Markdown formats", None, None),
        ("", None, None),
        ("NEED HELP?", subtitle_font, instruction_fill),
        ("• See the '📊 Sample Data' sheet for examples", None, None),
        ("• Refer to the Input_Data_Specification.docx for detailed requirements", None, None),
        ("• Contact your system administrator for technical support", None, None),
    ]

    for i, (text, font, fill) in enumerate(instructions, 1):
        cell = ws_instructions[f'A{i}']
        cell.value = text
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Sheet 2: Data Entry
    ws_data = wb.create_sheet("✏️ Data Entry", 1)

    # Set up columns
    columns = [
        ('id', 'A', 10, optional_fill),
        ('response', 'B', 60, required_fill),
        ('respondent_id', 'C', 15, optional_fill),
        ('timestamp', 'D', 20, optional_fill),
        ('topic', 'E', 20, optional_fill),
    ]

    # Add headers with formatting
    for col_name, col_letter, width, fill in columns:
        cell = ws_data[f'{col_letter}1']
        cell.value = col_name
        cell.font = header_font
        cell.fill = fill if col_name != 'response' else required_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws_data.column_dimensions[col_letter].width = width

    # Add explanation row
    explanations = [
        'Optional',
        'REQUIRED - Your text responses',
        'Optional',
        'Optional (YYYY-MM-DD)',
        'Optional'
    ]

    for i, (col_letter, explanation) in enumerate(zip(['A', 'B', 'C', 'D', 'E'], explanations)):
        cell = ws_data[f'{col_letter}2']
        cell.value = explanation
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add some example rows with light formatting
    ws_data.row_dimensions[2].height = 25

    # Freeze panes
    ws_data.freeze_panes = 'A3'

    # Sheet 3: Sample Data
    ws_sample = wb.create_sheet("📊 Sample Data", 2)

    # Sample data headers
    for col_name, col_letter, width, fill in columns:
        cell = ws_sample[f'{col_letter}1']
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws_sample.column_dimensions[col_letter].width = width

    # Sample data
    sample_data = [
        (1, "I love the flexibility of remote work", "R001", "2024-01-01", "Remote Work"),
        (2, "Better work-life balance is crucial", "R002", "2024-01-02", "Work-Life Balance"),
        (3, "Communication challenges with team members", "R003", "2024-01-03", "Communication"),
        (4, "Increased productivity at home", "R004", "2024-01-04", "Productivity"),
        (5, "Missing social interactions with colleagues", "R005", "2024-01-05", "Social Connection"),
        (6, "Technology issues affect my work", "R006", "2024-01-06", "Technology"),
        (7, "More time for family and personal activities", "R007", "2024-01-07", "Work-Life Balance"),
        (8, "Difficulty separating work and personal life", "R008", "2024-01-08", "Work-Life Balance"),
        (9, "Cost savings from not commuting", "R009", "2024-01-09", "Financial"),
        (10, "Feeling isolated from the team", "R010", "2024-01-10", "Social Connection"),
        (11, "Flexible schedule allows better time management", "R011", "2024-01-11", "Flexibility"),
        (12, "Video call fatigue is real", "R012", "2024-01-12", "Technology"),
        (13, "Can focus better without office distractions", "R013", "2024-01-13", "Productivity"),
        (14, "Miss casual conversations at the office", "R014", "2024-01-14", "Social Connection"),
        (15, "Home office setup improves comfort", "R015", "2024-01-15", "Work Environment"),
    ]

    for i, row_data in enumerate(sample_data, 2):
        for j, value in enumerate(row_data, 1):
            cell = ws_sample.cell(row=i, column=j)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Freeze panes
    ws_sample.freeze_panes = 'A2'

    # Save workbook
    wb.save(output_path)
    print(f"✓ Template created successfully: {output_path}")


if __name__ == "__main__":
    output_file = "/home/user/JC-OE-Coding/documentation/input_data_template.xlsx"
    create_data_template(output_file)
